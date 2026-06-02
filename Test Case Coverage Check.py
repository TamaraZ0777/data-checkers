import json
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


FILE_PATH = "Data/TC Data Check.xlsx"

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)


def create_backup(file_path):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    backup_path = file_path.replace(
        ".xlsx",
        f"_backup_{timestamp}.xlsx"
    )

    shutil.copy2(file_path, backup_path)

    print(f"Backup created: {backup_path}")


def parse_options(value):
    """
    Returns:
        options_list,
        source_format ('json' or 'text')
    """

    if value is None:
        return [], "text"

    text = str(value).strip()

    if not text:
        return [], "text"

    # Try JSON first
    try:
        parsed = json.loads(text)

        if isinstance(parsed, list):
            return [str(x).strip() for x in parsed], "json"

    except Exception:
        pass

    # Fallback: comma-separated text
    return (
        [x.strip() for x in text.split(",") if x.strip()],
        "text"
    )


def format_options(options, output_format):
    if output_format == "json":
        return json.dumps(
            options,
            ensure_ascii=False
        )

    return ", ".join(options)


def find_last_test_column(ws):

    last_test_col = None
    last_test_header = None

    for col in range(1, ws.max_column + 1):

        header = ws.cell(1, col).value

        if (
            isinstance(header, str)
            and header.startswith("test_")
        ):
            last_test_col = col
            last_test_header = header

    if last_test_col is None:
        raise Exception("No test_XX column found.")

    return last_test_col, last_test_header


def main():

    create_backup(FILE_PATH)

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    headers = {
        ws.cell(1, col).value: col
        for col in range(1, ws.max_column + 1)
    }

    if "answerType" not in headers:
        raise Exception("Column 'answerType' not found.")

    answer_type_col = headers["answerType"]

    test_col, test_header = find_last_test_column(ws)

    suffix = test_header.replace("test_", "")
    new_header = f"options_{suffix}"

    ws.insert_cols(test_col + 1)

    new_col = test_col + 1

    ws.cell(1, new_col).value = new_header

    source_options_col = test_col - 1

    for row in range(2, ws.max_row + 1):

        answer_type = ws.cell(
            row,
            answer_type_col
        ).value

        answer_type = (
            str(answer_type).strip().lower()
            if answer_type
            else ""
        )

        test_value = ws.cell(
            row,
            test_col
        ).value

        result_cell = ws.cell(
            row,
            new_col
        )

        # -------------------------
        # TEXT
        # -------------------------

        if answer_type == "text":

            if (
                test_value is None
                or str(test_value).strip() == ""
            ):
                result_cell.fill = GREEN_FILL

            continue

        # -------------------------
        # OPTIONS
        # -------------------------

        if answer_type not in ("option", "options"):
            continue

        source_value = ws.cell(
            row,
            source_options_col
        ).value

        options_list, source_format = parse_options(
            source_value
        )

        if (
            test_value is None
            or str(test_value).strip() == ""
        ):

            result_cell.value = format_options(
                options_list,
                source_format
            )

            result_cell.fill = GREEN_FILL

            continue

        selected_options = [
            item.strip()
            for item in str(test_value).split(",")
            if item.strip()
        ]

        remaining_options = options_list.copy()

        missing_options = []

        for option in selected_options:

            if option in remaining_options:
                remaining_options.remove(option)
            else:
                missing_options.append(option)

        result_text = format_options(
            remaining_options,
            source_format
        )

        if missing_options:

            result_text += (
                "\n\nNew option(s) not found "
                "in previous options list: "
                + ", ".join(missing_options)
            )

        result_cell.value = result_text

    wb.save(FILE_PATH)

    print(
        f"Processing completed.\n"
        f"Updated file: {FILE_PATH}"
    )


if __name__ == "__main__":
    main()
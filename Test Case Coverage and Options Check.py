import json
import re
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


FILE_PATH = "Data/TC Data Check.xlsx"


GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)

ORANGE_FILL = PatternFill(
    fill_type="solid",
    start_color="FCE4D6",
    end_color="FCE4D6"
)


def create_backup(file_path):

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = file_path.replace(
        ".xlsx",
        f"_backup_{timestamp}.xlsx"
    )

    shutil.copy2(file_path, backup_path)

    print(f"Backup created: {backup_path}")


def get_column_by_header(ws, header_name):

    for col in range(1, ws.max_column + 1):

        if ws.cell(1, col).value == header_name:
            return col

    return None


def find_next_test_to_process(ws):

    test_numbers = set()
    options_numbers = set()

    for col in range(1, ws.max_column + 1):

        header = ws.cell(1, col).value

        if not isinstance(header, str):
            continue

        match = re.match(r"test_(\d+)$", header)
        if match:
            test_numbers.add(int(match.group(1)))
            continue

        match = re.match(r"options_(\d+)$", header)
        if match:
            options_numbers.add(int(match.group(1)))

    pending = sorted(test_numbers - options_numbers)

    if not pending:
        return None

    return pending[0]

def parse_options(value):

    if value is None:
        return [], "text"

    text = str(value).strip()

    if not text:
        return [], "text"

    try:

        parsed = json.loads(text)

        if isinstance(parsed, list):

            return (
                [
                    str(x).strip()
                    for x in parsed
                ],
                "json"
            )

    except Exception:
        pass

    return (
        [
            x.strip()
            for x in text.split(",")
            if x.strip()
        ],
        "text"
    )


def format_options(options, output_format):

    if output_format == "json":

        return json.dumps(
            options,
            ensure_ascii=False
        )

    return ", ".join(options)


def row_has_previous_coverage(
    ws,
    row,
    current_test_num
):

    for test_num in range(
        1,
        current_test_num
    ):

        col = get_column_by_header(
            ws,
            f"test_{test_num:02d}"
        )

        if col is None:
            continue

        value = ws.cell(
            row,
            col
        ).value

        if (
            value is not None
            and str(value).strip()
        ):
            return True

    return False


def main():

    create_backup(FILE_PATH)

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    answer_type_col = get_column_by_header(
        ws,
        "answerType"
    )

    if answer_type_col is None:

        raise Exception(
            "answerType column not found."
        )

    test_num = find_next_test_to_process(ws)

    if test_num is None:

        print(
            "No unprocessed test columns found."
        )

        return

    suffix = f"{test_num:02d}"

    test_header = f"test_{suffix}"

    test_col = get_column_by_header(
        ws,
        test_header
    )

    if test_col is None:

        raise Exception(
            f"{test_header} not found."
        )

    
    options_header = (
        f"options_{suffix}"
    )

    new_option_header = (
        f"new_option_found_{suffix}"
    )

    existing_col = get_column_by_header(
        ws,
        options_header
    )

    if existing_col is not None:

        print(
            f"{options_header} already exists."
        )

        print("Nothing to do.")

        return

    ws.insert_cols(
        test_col + 1,
        amount=2
    )

    options_col = test_col + 1
    new_option_col = test_col + 2

    ws.cell(
        1,
        options_col
    ).value = options_header

    ws.cell(
        1,
        new_option_col
    ).value = new_option_header

    # Original options column
    original_options_col = (
        get_column_by_header(
            ws,
            "options"
        )
    )

    if original_options_col is None:

        raise Exception(
            "Original options column not found."
        )

    # Coverage source column
    if test_num == 1:

        source_options_col = (
            original_options_col
        )

    else:

        source_options_col = (
            get_column_by_header(
                ws,
                f"options_{test_num - 1:02d}"
            )
        )

        if source_options_col is None:

            raise Exception(
                f"options_{test_num - 1:02d} "
                "not found."
            )

    for row in range(
        2,
        ws.max_row + 1
    ):

        answer_type = ws.cell(
            row,
            answer_type_col
        ).value

        answer_type = (
            str(answer_type)
            .strip()
            .lower()
            if answer_type
            else ""
        )

        test_value = ws.cell(
            row,
            test_col
        ).value

        options_cell = ws.cell(
            row,
            options_col
        )

        new_option_cell = ws.cell(
            row,
            new_option_col
        )

        has_previous_coverage = (
            row_has_previous_coverage(
                ws,
                row,
                test_num
            )
        )

        # -------------------------------------------------
        # TEXT
        # -------------------------------------------------

        if answer_type == "text":

            if (
                (
                    test_value is None
                    or not str(
                        test_value
                    ).strip()
                )
                and not has_previous_coverage
            ):

                options_cell.fill = (
                    GREEN_FILL
                )

            continue

        # -------------------------------------------------
        # OPTIONS
        # -------------------------------------------------

        if answer_type not in (
            "option",
            "options"
        ):
            continue

        source_value = ws.cell(
            row,
            source_options_col
        ).value

        (
            remaining_options,
            source_format
        ) = parse_options(
            source_value
        )

        (
            original_options,
            _
        ) = parse_options(
            ws.cell(
                row,
                original_options_col
            ).value
        )

        if (
            test_value is None
            or not str(
                test_value
            ).strip()
        ):

            options_cell.value = (
                format_options(
                    remaining_options,
                    source_format
                )
            )

            if not has_previous_coverage:

                options_cell.fill = (
                    GREEN_FILL
                )

            continue

        selected_options = [

            item.strip()

            for item in str(
                test_value
            ).split(",")

            if item.strip()
        ]

        missing_options = []

        updated_remaining = (
            remaining_options.copy()
        )

        for option in selected_options:

            if (
                option
                not in original_options
            ):

                missing_options.append(
                    option
                )

                continue

            if option in updated_remaining:

                updated_remaining.remove(
                    option
                )

        options_cell.value = (
            format_options(
                updated_remaining,
                source_format
            )
        )

        if missing_options:

            new_option_cell.value = (
                ", ".join(
                    missing_options
                )
            )

            new_option_cell.fill = (
                ORANGE_FILL
            )

    wb.save(FILE_PATH)

    print(
        f"Processing completed.\n"
        f"Updated file: {FILE_PATH}"
    )


if __name__ == "__main__":
    main()
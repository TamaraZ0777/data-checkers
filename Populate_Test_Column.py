import os
import re
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


TC_FILE = "Data/TC Data Check.xlsx"
DATA_FILES_FOLDER = "Data Files"


# --------------------------------------------------
# Backup
# --------------------------------------------------

def create_backup(file_path):

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = file_path.replace(
        ".xlsx",
        f"_backup_{timestamp}.xlsx"
    )

    shutil.copy2(file_path, backup_path)

    print(f"Backup created: {backup_path}")


# --------------------------------------------------
# Helpers
# --------------------------------------------------

def get_column_by_header(ws, header_name):

    for col in range(1, ws.max_column + 1):

        if ws.cell(1, col).value == header_name:
            return col

    return None


# --------------------------------------------------
# Find next test number
# --------------------------------------------------

def get_next_test_number(ws):

    test_numbers = []

    for col in range(1, ws.max_column + 1):

        header = ws.cell(1, col).value

        if not isinstance(header, str):
            continue

        match = re.match(
            r"test_(\d+)$",
            header
        )

        if match:

            test_numbers.append(
                int(match.group(1))
            )

    if not test_numbers:
        return 1

    test_numbers.sort()

    expected = 1

    for number in test_numbers:

        if number != expected:

            raise Exception(
                f"Gap detected.\n"
                f"Missing test_{expected:02d}"
            )

        expected += 1

    return expected


# --------------------------------------------------
# Find source file
# --------------------------------------------------

def get_file_number(filename):

    match = re.search(
        r"(\d+)\s+Export\.xlsx$",
        filename,
        re.IGNORECASE
    )

    if match:
        return int(match.group(1))

    return None


def find_source_file(expected_number):

    matches = []

    for file in os.listdir(DATA_FILES_FOLDER):

        if not file.lower().endswith(".xlsx"):
            continue

        number = get_file_number(file)

        if number == expected_number:

            matches.append(
                os.path.join(
                    DATA_FILES_FOLDER,
                    file
                )
            )

    if not matches:

        raise Exception(
            f"No source file found for "
            f"test_{expected_number:02d}"
        )

    if len(matches) > 1:

        raise Exception(
            "Multiple matching files found:\n"
            + "\n".join(matches)
        )

    return matches[0]


# --------------------------------------------------
# Main
# --------------------------------------------------

def main():

    create_backup(TC_FILE)

    wb = load_workbook(TC_FILE)
    ws = wb.active

    next_test_num = (
        get_next_test_number(ws)
    )

    print(
        f"Next test number: "
        f"{next_test_num:02d}"
    )

    source_file = find_source_file(
        next_test_num
    )

    print(
        f"Source file found:\n"
        f"{source_file}"
    )

    df = pd.read_excel(source_file)

    if "...Text..." not in df.columns:

        raise Exception(
            "'...Text...' column "
            "not found."
        )

    tc_rows = ws.max_row - 1
    source_rows = len(df)

    if tc_rows != source_rows:

        raise Exception(
            "Row count mismatch.\n\n"
            f"TC Data Check: "
            f"{tc_rows}\n"
            f"Source file: "
            f"{source_rows}"
        )

    new_header = (
        f"test_{next_test_num:02d}"
    )

    if get_column_by_header(
        ws,
        new_header
    ):

        raise Exception(
            f"{new_header} already exists."
        )

    new_col = ws.max_column + 1

    ws.cell(
        1,
        new_col
    ).value = new_header

    for excel_row, value in enumerate(
        df["...Text..."],
        start=2
    ):

        if pd.isna(value):

            value = ""

        else:

            value = str(value).strip()

        ws.cell(
            excel_row,
            new_col
        ).value = value

    wb.save(TC_FILE)

    print(
        f"\nSuccessfully added "
        f"{new_header}"
    )


if __name__ == "__main__":
    main()